"""
Enhanced Excel export utilities for comprehensive portfolio analysis reports.
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import io

# Try to import openpyxl, fall back to pandas-only solution if not available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, Reference, BarChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExportManager:
    """Manages comprehensive Excel export functionality for portfolio analysis."""
    
    def __init__(self):
        self.use_openpyxl = OPENPYXL_AVAILABLE
        if self.use_openpyxl:
            self.workbook = Workbook()
            # Remove default sheet
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
            self._setup_styles()
        else:
            self.workbook = None
    
    def _setup_styles(self):
        """Setup common styles for the workbook."""
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.fill = PatternFill("solid", fgColor="366092")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Data style
        data_style = NamedStyle(name="data")
        data_style.alignment = Alignment(horizontal="center", vertical="center")
        data_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Percentage style
        pct_style = NamedStyle(name="percentage")
        pct_style.number_format = "0.00%"
        pct_style.alignment = Alignment(horizontal="center", vertical="center")
        pct_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Currency style
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = "$#,##0.00"
        currency_style.alignment = Alignment(horizontal="center", vertical="center")
        currency_style.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        try:
            self.workbook.add_named_style(header_style)
            self.workbook.add_named_style(data_style)
            self.workbook.add_named_style(pct_style)
            self.workbook.add_named_style(currency_style)
        except ValueError:
            # Styles already exist
            pass
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths for a worksheet."""
        from openpyxl.utils import get_column_letter
        
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    # Skip merged cells
                    if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                        continue
                    
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 4, 40)  # More generous width
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_comprehensive_report(self, analysis_results: Dict, sip_amount: float = 0) -> Tuple[bytes, str, str]:
        """Create comprehensive Excel report with multiple sheets containing portfolio insights.
        
        Returns:
            Tuple of (file_data, filename, mime_type)
        """
        
        if self.use_openpyxl:
            return self._create_enhanced_openpyxl_report(analysis_results, sip_amount)
        else:
            return self._create_enhanced_pandas_report(analysis_results, sip_amount)
    
    def _create_enhanced_openpyxl_report(self, analysis_results, sip_amount):
        """Create comprehensive Excel report using openpyxl for advanced formatting."""
        
        # Sheet 1: Executive Summary
        self._create_executive_summary_sheet(analysis_results, sip_amount)
        
        # Sheet 2: Strategy Performance Comparison
        self._create_strategy_comparison_sheet(analysis_results)
        
        # Sheet 3: Risk Analysis
        self._create_risk_analysis_sheet(analysis_results)
        
        # Sheet 4: Portfolio Composition
        self._create_portfolio_composition_sheet(analysis_results)
        
        # Sheet 5: Monthly Returns Analysis
        self._create_returns_analysis_sheet(analysis_results)
        
        # Sheet 6: Asset Correlation Matrix
        self._create_correlation_matrix_sheet(analysis_results)
        
        # Sheet 7: Portfolio Growth Timeline
        self._create_growth_timeline_sheet(analysis_results, sip_amount)
        
        # Generate filename and save to bytes
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Portfolio_Analysis_Report_{timestamp}.xlsx"
        
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    def _create_executive_summary_sheet(self, analysis_results, sip_amount):
        """Create executive summary sheet with key insights."""
        ws = self.workbook.create_sheet(title="Executive Summary")
        
        # Title
        ws['A1'] = "Portfolio Analysis - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Analysis metadata
        ws['A3'] = "Analysis Date:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws['A4'] = "Monthly Investment:"
        ws['B4'] = sip_amount
        ws['B4'].style = "currency"
        
        # Get performance summary
        performance_summary = analysis_results.get('performance_summary', {})
        
        # Key metrics summary
        row = 6
        ws[f'A{row}'] = "Strategy Performance Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Headers
        headers = ['Strategy', 'XIRR', 'Total Return', 'Volatility', 'Sharpe Ratio', 'Max Drawdown']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.style = "header"
        row += 1
        
        # Data rows
        for strategy_name, metrics in performance_summary.items():
            ws.cell(row=row, column=1, value=strategy_name).style = "data"
            ws.cell(row=row, column=2, value=metrics.get('XIRR', 0)).style = "percentage"
            ws.cell(row=row, column=3, value=metrics.get('Total Return', 0)).style = "percentage"
            ws.cell(row=row, column=4, value=metrics.get('Volatility', 0)).style = "percentage"
            ws.cell(row=row, column=5, value=metrics.get('Sharpe Ratio', 0)).style = "data"
            ws.cell(row=row, column=6, value=metrics.get('Max Drawdown', 0)).style = "percentage"
            row += 1
        
        # Best performing strategy and recommendations
        if performance_summary:
            best_strategy = max(performance_summary.items(), key=lambda x: x[1].get('XIRR', 0))
            row += 2
            ws[f'A{row}'] = "🏆 Best Performing Strategy:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = best_strategy[0]
            ws[f'C{row}'] = f"{best_strategy[1].get('XIRR', 0):.2%} XIRR"
            
            # Add key insights
            row += 2
            ws[f'A{row}'] = "📊 Key Portfolio Insights"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            config = analysis_results.get('config', {})
            tickers = config.get('tickers', [])
            
            ws[f'A{row}'] = f"• Total Assets Selected: {len(tickers)}"
            row += 1
            ws[f'A{row}'] = f"• Investment Period: {len(performance_summary)} strategies analyzed"
            row += 1
            
            # Risk assessment
            best_volatility = best_strategy[1].get('Volatility', 0)
            if best_volatility < 0.15:
                risk_level = "Conservative (Low Risk)"
            elif best_volatility < 0.25:
                risk_level = "Moderate (Medium Risk)"
            else:
                risk_level = "Aggressive (High Risk)"
            
            ws[f'A{row}'] = f"• Risk Profile: {risk_level}"
            row += 1
            
            # Expected returns
            expected_annual = best_strategy[1].get('XIRR', 0)
            expected_monthly = sip_amount
            expected_annual_investment = expected_monthly * 12
            projected_value = expected_annual_investment * (1 + expected_annual)
            
            ws[f'A{row}'] = f"• Projected Annual Investment: ${expected_annual_investment:,.0f}"
            row += 1
            ws[f'A{row}'] = f"• Projected Year 1 Value: ${projected_value:,.0f}"
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_strategy_comparison_sheet(self, analysis_results):
        """Create detailed strategy comparison sheet."""
        ws = self.workbook.create_sheet(title="Strategy Comparison")
        
        performance_summary = analysis_results.get('performance_summary', {})
        
        # Title
        ws['A1'] = "Detailed Strategy Performance Comparison"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:H1')
        
        # Headers
        row = 3
        headers = ['Strategy Name', 'XIRR', 'Total Return', 'Final Value', 'Total Contributions', 
                  'Volatility', 'Sharpe Ratio', 'Max Drawdown']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.style = "header"
        row += 1
        
        # Data rows
        for strategy_name, metrics in performance_summary.items():
            ws.cell(row=row, column=1, value=strategy_name).style = "data"
            ws.cell(row=row, column=2, value=metrics.get('XIRR', 0)).style = "percentage"
            ws.cell(row=row, column=3, value=metrics.get('Total Return', 0)).style = "percentage"
            ws.cell(row=row, column=4, value=metrics.get('Final Value', 0)).style = "currency"
            ws.cell(row=row, column=5, value=metrics.get('Total Contributions', 0)).style = "currency"
            ws.cell(row=row, column=6, value=metrics.get('Volatility', 0)).style = "percentage"
            ws.cell(row=row, column=7, value=metrics.get('Sharpe Ratio', 0)).style = "data"
            ws.cell(row=row, column=8, value=metrics.get('Max Drawdown', 0)).style = "percentage"
        row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_risk_analysis_sheet(self, analysis_results):
        """Create risk analysis sheet."""
        ws = self.workbook.create_sheet(title="Risk Analysis")
        
        ws['A1'] = "Portfolio Risk Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        performance_summary = analysis_results.get('performance_summary', {})
        
        # Risk metrics comparison
        row = 3
        ws[f'A{row}'] = "Risk Metrics by Strategy"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 2
        
        headers = ['Strategy', 'Volatility', 'Max Drawdown', 'Sharpe Ratio', 'Risk Level']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.style = "header"
        row += 1
            
        for strategy_name, metrics in performance_summary.items():
            volatility = metrics.get('Volatility', 0)
            # Determine risk level based on volatility
            if volatility < 0.1:
                risk_level = "Low"
            elif volatility < 0.2:
                risk_level = "Medium"
            else:
                risk_level = "High"
            
            ws.cell(row=row, column=1, value=strategy_name).style = "data"
            ws.cell(row=row, column=2, value=volatility).style = "percentage"
            ws.cell(row=row, column=3, value=metrics.get('Max Drawdown', 0)).style = "percentage"
            ws.cell(row=row, column=4, value=metrics.get('Sharpe Ratio', 0)).style = "data"
            ws.cell(row=row, column=5, value=risk_level).style = "data"
            row += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_portfolio_composition_sheet(self, analysis_results):
        """Create portfolio composition analysis sheet."""
        ws = self.workbook.create_sheet(title="Portfolio Composition")
        
        ws['A1'] = "Portfolio Asset Composition & Allocation"
        ws['A1'].font = Font(size=14, bold=True)
        
        config = analysis_results.get('config', {})
        tickers = config.get('tickers', [])
        
        # Get the best strategy's allocation if available
        performance_summary = analysis_results.get('performance_summary', {})
        results = analysis_results.get('results', {})
        
        best_strategy_name = None
        if performance_summary:
            best_strategy = max(performance_summary.items(), key=lambda x: x[1].get('XIRR', 0))
            best_strategy_name = best_strategy[0]
        
        # Asset list with allocations
        row = 3
        ws[f'A{row}'] = f"Asset Allocation - {best_strategy_name if best_strategy_name else 'Equal Weight'}"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 2
        
        ws.cell(row=row, column=1, value="Asset Symbol").style = "header"
        ws.cell(row=row, column=2, value="Asset Type").style = "header"
        ws.cell(row=row, column=3, value="Allocation %").style = "header"
        ws.cell(row=row, column=4, value="Monthly Investment").style = "header"
        row += 1
        
        # Asset categories for classification
        asset_categories = {
            "US Equity ETFs": ["VOO", "SPY", "VTI", "QQQ", "IWM", "VTV", "VUG", "VB", "VO"],
            "International ETFs": ["VEA", "VWO", "IEFA", "IEMG", "VGK", "VPL", "VSS"],
            "Bonds & Fixed Income": ["BND", "VGIT", "VGLT", "VTEB", "LQD", "HYG", "TIP"],
            "Cryptocurrency": ["BTC", "ETH"],
            "Sector ETFs": ["XLK", "XLF", "XLV", "XLE", "XLI", "XLP", "XLY", "XLU", "XLRE"],
            "Commodities & REITs": ["GLD", "SLV", "USO", "VNQ", "VNQI", "DBA", "DJP"]
        }
        
        def get_asset_type(ticker):
            for category, tickers_list in asset_categories.items():
                if ticker in tickers_list:
                    return category
            return "Individual Stock"
        
        # Calculate actual allocations from best strategy's final portfolio composition
        sip_amount = config.get('sip_amount', 0)
        results = analysis_results.get('results', {})
        
        # Use EXACT same logic as dashboard to get optimal weights
        allocation_weights = {}
        allocation_method = "Equal Weight (Fallback)"
        
        try:
            # Import here to avoid circular imports (same as dashboard)
            from models.sip_strategy import SIPStrategyAnalyzer
            
            # Create analyzer instance (same as dashboard)
            analyzer = SIPStrategyAnalyzer(config)
            analyzer.available_tickers = config['tickers']  # Use config tickers, not filtered tickers
            
            # Get data for optimization (same as dashboard)
            data = analysis_results.get('data')
            
            if data is not None and best_strategy_name:
                # Determine optimization method from strategy name (EXACT same logic as dashboard)
                if "Max Sharpe" in best_strategy_name:
                    method = "max_sharpe"
                elif "Min Variance" in best_strategy_name:
                    method = "min_variance"
                elif "Equal Weight" in best_strategy_name:
                    method = "equal_weight"
                else:
                    # Default to max_sharpe if method unclear (same as dashboard)
                    method = "max_sharpe"
                
                if method == "equal_weight":
                    # Equal weight allocation (same as dashboard)
                    num_assets = len(config['tickers'])
                    allocation_weights = {ticker: 1.0/num_assets for ticker in config['tickers']}
                    allocation_method = "Equal Weight Strategy"
                else:
                    # Get optimal weights from MPT optimization (EXACT same call as dashboard)
                    optimal_weights = analyzer.get_optimal_allocation_weights(data, method)
                    
                    if optimal_weights:
                        allocation_weights = optimal_weights
                        allocation_method = f"MPT Optimized ({method.replace('_', ' ').title()}) - Same as Dashboard"
                    else:
                        # Fallback to equal weight (same as dashboard)
                        allocation_weights = {ticker: 1.0/len(config['tickers']) for ticker in config['tickers']}
                        allocation_method = "Equal Weight (Optimization Failed)"
            else:
                # Fallback to equal weight (same as dashboard)
                allocation_weights = {ticker: 1.0/len(config['tickers']) for ticker in config['tickers']}
                allocation_method = "Equal Weight (No Data)"
                
        except Exception as e:
            # Fallback to equal weight if any error (same as dashboard)
            allocation_weights = {ticker: 1.0/len(config['tickers']) for ticker in config['tickers']}
            allocation_method = f"Equal Weight (Error: {str(e)[:30]}...)"
        
        # Use the same tickers as in the allocation_weights (which comes from config['tickers'])
        display_tickers = list(allocation_weights.keys()) if allocation_weights else tickers
        
        # Add diagnostic information
        row += 1
        ws[f'A{row}'] = f"Method: {allocation_method}"
        ws[f'A{row}'].font = Font(size=9, italic=True)
        row += 1
        
        # Add debug info
        ws[f'A{row}'] = f"Config tickers: {len(config.get('tickers', []))}, Display tickers: {len(display_tickers)}"
        ws[f'A{row}'].font = Font(size=8, italic=True)
        row += 1
        
        for ticker in display_tickers:
            allocation_pct = allocation_weights.get(ticker, 1.0 / len(display_tickers))
            monthly_investment = allocation_pct * sip_amount
            
            ws.cell(row=row, column=1, value=ticker).style = "data"
            ws.cell(row=row, column=2, value=get_asset_type(ticker)).style = "data"
            ws.cell(row=row, column=3, value=allocation_pct).style = "percentage"
            ws.cell(row=row, column=4, value=monthly_investment).style = "currency"
            row += 1
        
        # Add total row
        row += 1
        ws.cell(row=row, column=2, value="TOTAL").style = "header"
        ws.cell(row=row, column=3, value=1.0).style = "percentage"
        ws.cell(row=row, column=4, value=sip_amount).style = "currency"
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_returns_analysis_sheet(self, analysis_results):
        """Create comprehensive returns analysis sheet."""
        ws = self.workbook.create_sheet(title="Returns Analysis")
        
        ws['A1'] = "Comprehensive Returns & Performance Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        performance_summary = analysis_results.get('performance_summary', {})
        if performance_summary:
            # Detailed performance metrics
            row = 3
            ws[f'A{row}'] = "Detailed Performance Metrics by Strategy"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 2
            
            headers = ['Strategy Name', 'Annualized Return (XIRR)', 'Total Return', 
                      'Final Portfolio Value', 'Total Invested', 'Absolute Gain/Loss',
                      'Risk-Adjusted Return (Sharpe)', 'Volatility', 'Max Drawdown']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.style = "header"
            row += 1
            
            for strategy_name, metrics in performance_summary.items():
                final_value = metrics.get('Final Value', 0)
                total_contributions = metrics.get('Total Contributions', 0)
                absolute_gain = final_value - total_contributions
                
                ws.cell(row=row, column=1, value=strategy_name).style = "data"
                ws.cell(row=row, column=2, value=metrics.get('XIRR', 0)).style = "percentage"
                ws.cell(row=row, column=3, value=metrics.get('Total Return', 0)).style = "percentage"
                ws.cell(row=row, column=4, value=final_value).style = "currency"
                ws.cell(row=row, column=5, value=total_contributions).style = "currency"
                ws.cell(row=row, column=6, value=absolute_gain).style = "currency"
                ws.cell(row=row, column=7, value=metrics.get('Sharpe Ratio', 0)).style = "data"
                ws.cell(row=row, column=8, value=metrics.get('Volatility', 0)).style = "percentage"
                ws.cell(row=row, column=9, value=metrics.get('Max Drawdown', 0)).style = "percentage"
                row += 1
            
            # Performance ranking
            row += 2
            ws[f'A{row}'] = "Strategy Rankings"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            # Sort strategies by XIRR
            sorted_strategies = sorted(performance_summary.items(), 
                                     key=lambda x: x[1].get('XIRR', 0), reverse=True)
            
            ws.cell(row=row, column=1, value="Rank").style = "header"
            ws.cell(row=row, column=2, value="Strategy").style = "header"
            ws.cell(row=row, column=3, value="XIRR").style = "header"
            ws.cell(row=row, column=4, value="Performance Grade").style = "header"
            row += 1
            
            for rank, (strategy_name, metrics) in enumerate(sorted_strategies, 1):
                xirr = metrics.get('XIRR', 0)
                
                # Assign performance grade
                if xirr >= 0.15:
                    grade = "Excellent"
                elif xirr >= 0.12:
                    grade = "Very Good"
                elif xirr >= 0.10:
                    grade = "Good"
                elif xirr >= 0.08:
                    grade = "Fair"
                else:
                    grade = "Poor"
                
                ws.cell(row=row, column=1, value=rank).style = "data"
                ws.cell(row=row, column=2, value=strategy_name).style = "data"
                ws.cell(row=row, column=3, value=xirr).style = "percentage"
                ws.cell(row=row, column=4, value=grade).style = "data"
        row += 1
    
    def _create_correlation_matrix_sheet(self, analysis_results):
        """Create asset correlation matrix sheet."""
        ws = self.workbook.create_sheet(title="Correlation Matrix")
        
        ws['A1'] = "Asset Correlation Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Get data for correlation analysis
        data = analysis_results.get('data')
        if data is not None and not data.empty:
            # Calculate correlation matrix
            correlation_matrix = data.corr()
            
            row = 3
            ws[f'A{row}'] = "Asset Correlation Matrix"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 2
            
            # Headers
            tickers = list(correlation_matrix.columns)
            ws.cell(row=row, column=1, value="Asset").style = "header"
            for col, ticker in enumerate(tickers, 2):
                ws.cell(row=row, column=col, value=ticker).style = "header"
            row += 1
            
            # Correlation data
            for i, ticker1 in enumerate(tickers):
                ws.cell(row=row, column=1, value=ticker1).style = "header"
                for j, ticker2 in enumerate(tickers):
                    corr_value = correlation_matrix.loc[ticker1, ticker2]
                    cell = ws.cell(row=row, column=j+2, value=corr_value)
                    cell.style = "data"
                    cell.number_format = "0.00"
                    
                    # Color coding for correlation strength
                    if corr_value > 0.7:
                        cell.fill = PatternFill("solid", fgColor="90EE90")  # Light green
                    elif corr_value < -0.7:
                        cell.fill = PatternFill("solid", fgColor="FFB6C1")  # Light red
                row += 1
        else:
            row = 3
            ws[f'A{row}'] = "Correlation analysis requires price data"
            ws[f'A{row}'].font = Font(italic=True)
    
    def _create_growth_timeline_sheet(self, analysis_results, sip_amount):
        """Create portfolio growth timeline sheet."""
        ws = self.workbook.create_sheet(title="Growth Timeline")
        
        ws['A1'] = "Portfolio Growth Timeline"
        ws['A1'].font = Font(size=14, bold=True)
        
        results = analysis_results.get('results', {})
        
        if results:
            # Get the best performing strategy for timeline
            performance_summary = analysis_results.get('performance_summary', {})
            if performance_summary:
                best_strategy = max(performance_summary.items(), key=lambda x: x[1].get('XIRR', 0))
                best_strategy_name = best_strategy[0]
                
                if best_strategy_name in results:
                    portfolio_values = results[best_strategy_name].get('portfolio_values', pd.Series())
                    contributions = results[best_strategy_name].get('monthly_contributions', pd.Series())
                    
                    row = 3
                    ws[f'A{row}'] = f"Growth Timeline - {best_strategy_name}"
                    ws[f'A{row}'].font = Font(size=12, bold=True)
                    row += 2
                    
                    # Headers
                    headers = ['Date', 'Portfolio Value', 'Monthly Contribution', 'Cumulative Contributions']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.style = "header"
                    row += 1
                    
                    # Generate proper monthly SIP dates for timeline
                    if len(portfolio_values) > 0:
                        start_date = portfolio_values.index[0]
                        end_date = portfolio_values.index[-1]
                        
                        # Generate monthly SIP dates (first business day of each month)
                        monthly_sip_dates = pd.date_range(start=start_date, end=end_date, freq='MS')
                        
                        cumulative_contributions = 0
                        
                        # Show up to 24 months to avoid too many rows
                        display_dates = monthly_sip_dates[:24] if len(monthly_sip_dates) > 24 else monthly_sip_dates
                        
                        for sip_date in display_dates:
                            # Find closest portfolio value date
                            closest_date = None
                            if sip_date in portfolio_values.index:
                                closest_date = sip_date
                            else:
                                # Find closest available date
                                available_dates = portfolio_values.index[portfolio_values.index >= sip_date]
                                if len(available_dates) > 0:
                                    closest_date = available_dates[0]
                                else:
                                    # Look backwards
                                    available_dates = portfolio_values.index[portfolio_values.index <= sip_date]
                                    if len(available_dates) > 0:
                                        closest_date = available_dates[-1]
                            
                            if closest_date is not None:
                                value = portfolio_values.loc[closest_date]
                                
                                # Check if this is an actual SIP date with contribution
                                contribution = sip_amount  # Default to SIP amount for monthly dates
                                if sip_date in contributions.index:
                                    actual_contribution = contributions.loc[sip_date]
                                    if not pd.isna(actual_contribution) and actual_contribution > 0:
                                        contribution = actual_contribution
                                
                                cumulative_contributions += contribution
                                
                                ws.cell(row=row, column=1, value=sip_date.strftime('%Y-%m-%d')).style = "data"
                                ws.cell(row=row, column=2, value=value).style = "currency"
                                ws.cell(row=row, column=3, value=contribution).style = "currency"
                                ws.cell(row=row, column=4, value=cumulative_contributions).style = "currency"
                                row += 1
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_enhanced_pandas_report(self, analysis_results, sip_amount):
        """Create Excel report using pandas (fallback when openpyxl not available)."""
        
        # Create a simple multi-sheet Excel file using pandas
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Executive Summary
            performance_summary = analysis_results.get('performance_summary', {})
            if performance_summary:
                summary_df = pd.DataFrame(performance_summary).T
                summary_df.to_excel(writer, sheet_name='Executive Summary', index=True)
            
            # Strategy Comparison
            if performance_summary:
                comparison_df = pd.DataFrame(performance_summary).T
                comparison_df.to_excel(writer, sheet_name='Strategy Comparison', index=True)
            
            # Portfolio Composition
            config = analysis_results.get('config', {})
            tickers = config.get('tickers', [])
            if tickers:
                composition_df = pd.DataFrame({'Assets': tickers})
                composition_df.to_excel(writer, sheet_name='Portfolio Composition', index=False)
            
            # Correlation Matrix
            data = analysis_results.get('data')
            if data is not None and not data.empty:
                correlation_matrix = data.corr()
                correlation_matrix.to_excel(writer, sheet_name='Correlation Matrix', index=True)
        
        buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Portfolio_Analysis_Report_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"